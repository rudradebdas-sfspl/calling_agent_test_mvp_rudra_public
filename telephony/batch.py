"""
batch.py — Sequential Batch Calling with Retry Logic

FLOW:
  Call 1 dispatch
      ↓ (30 seconds wait)
  Call 2 try → if failed → retry → retry → retry... until SUCCESS
      ↓ (30 seconds wait after success)
  Call 3 try → if failed → retry → retry... until SUCCESS
      ↓
  ... and so on until list ends

CHANGED: Now parses optional name/gender columns from Excel.
"""

import asyncio
import io
import logging
import re
import uuid
from typing import Optional

import openpyxl

from app.telephony.services import dispatch_single_call

logger = logging.getLogger("telephony.batch")

# Seconds to wait AFTER dispatching a call before trying the next one
WAIT_AFTER_CALL = 30.0

# Seconds to wait between retry attempts when a call fails
RETRY_INTERVAL = 10.0

# Phone number validation
PHONE_REGEX = re.compile(r"^\+?\d{10,15}$")

# In-memory progress tracker
batch_progress: dict[str, dict] = {}


def validate_phone_number(number: str) -> Optional[str]:
    cleaned = number.strip().replace(" ", "").replace("-", "")
    if not cleaned.startswith("+"):
        if cleaned.startswith("0"):
            cleaned = "+91" + cleaned[1:]
        elif len(cleaned) == 10:
            cleaned = "+91" + cleaned
    if PHONE_REGEX.match(cleaned):
        return cleaned
    return None


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse Excel file. Returns list of dicts with phone, name, gender."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}")

    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active worksheet")

    header_row = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    phone_col_idx = None
    name_col_idx = None
    gender_col_idx = None

    for idx, header in enumerate(header_row):
        if header in ("phone_number", "phone", "number", "mobile", "phonenumber"):
            phone_col_idx = idx
        elif header in ("name", "callee_name", "contact_name"):
            name_col_idx = idx
        elif header in ("gender", "sex"):
            gender_col_idx = idx

    if phone_col_idx is None:
        raise ValueError(
            "Excel file must have a column named 'phone_number' (or 'phone', 'number', 'mobile'). "
            f"Found: {header_row}"
        )

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if phone_col_idx < len(row) and row[phone_col_idx] is not None:
            val = str(row[phone_col_idx]).strip()
            if val:
                entry = {"phone": val}
                if name_col_idx is not None and name_col_idx < len(row) and row[name_col_idx]:
                    entry["name"] = str(row[name_col_idx]).strip()
                if gender_col_idx is not None and gender_col_idx < len(row) and row[gender_col_idx]:
                    entry["gender"] = str(row[gender_col_idx]).strip().lower()
                contacts.append(entry)

    wb.close()
    return contacts


def prepare_batch(file_bytes: bytes) -> dict:
    """Parse + validate Excel. Returns immediately. Actual calling happens in background."""
    batch_id = f"batch_{uuid.uuid4().hex[:12]}"

    raw_contacts = parse_excel(file_bytes)
    if not raw_contacts:
        return {"batch_id": batch_id, "total": 0, "validated": [], "errors": []}

    validated = []
    errors = []
    for entry in raw_contacts:
        cleaned = validate_phone_number(entry["phone"])
        if cleaned:
            validated.append({
                "phone": cleaned,
                "name": entry.get("name"),
                "gender": entry.get("gender"),
            })
        else:
            errors.append({"phone_number": entry["phone"], "error": "Invalid phone number format"})

    batch_progress[batch_id] = {
        "batch_id":      batch_id,
        "total":         len(validated),
        "current_index": 0,
        "current_number": None,
        "dispatched":    0,
        "failed_total":  len(errors),
        "retry_count":   0,
        "status":        "queued",
        "errors":        errors,
    }

    logger.info(f"[BATCH PREPARED] {batch_id}: {len(validated)} valid numbers")

    return {
        "batch_id":  batch_id,
        "total":     len(raw_contacts),
        "validated": validated,
        "errors":    errors,
    }


async def dispatch_with_retry(
    contact: dict,
    batch_id: str,
    progress: dict,
    agent_id: Optional[int] = None,
) -> bool:
    """
    Keep retrying a single call until SIP trunk accepts it (success=True).
    """
    number = contact["phone"]
    attempt = 0
    while True:
        attempt += 1
        progress["retry_count"] = attempt - 1

        logger.info(
            f"[BATCH] {batch_id}: Attempting call → {number} "
            f"(attempt #{attempt})"
        )

        result = await dispatch_single_call(
            phone_number=number,
            batch_id=batch_id,
            name=contact.get("name"),
            gender=contact.get("gender"),
            agent_id=agent_id,
        )

        if result["success"]:
            logger.info(
                f"[BATCH] {batch_id}: ✅ Call SUCCESS → {number} "
                f"on attempt #{attempt} | SID: {result.get('call_sid')}"
            )
            progress["retry_count"] = 0
            return True
        else:
            err = result.get("error", "Unknown error")
            logger.warning(
                f"[BATCH] {batch_id}: ❌ Call FAILED → {number} "
                f"(attempt #{attempt}) | {err} | Retrying in {RETRY_INTERVAL}s..."
            )
            await asyncio.sleep(RETRY_INTERVAL)


async def run_batch_background(batch_id: str, validated: list, agent_id: Optional[int] = None) -> None:
    """
    Background task: Dispatch calls sequentially with retry-until-success logic.
    validated is now a list of dicts: [{"phone": ..., "name": ..., "gender": ...}, ...]
    """
    progress = batch_progress.setdefault(batch_id, {})
    progress["status"] = "running"
    dispatched = 0

    logger.info(f"[BATCH START] {batch_id}: {len(validated)} calls to process sequentially")

    for i, contact in enumerate(validated):
        number = contact["phone"]
        progress["current_index"] = i + 1
        progress["current_number"] = number

        logger.info(
            f"[BATCH] {batch_id}: ── Call {i+1}/{len(validated)} → {number} ──"
        )

        await dispatch_with_retry(contact, batch_id, progress, agent_id=agent_id)
        dispatched += 1
        progress["dispatched"] = dispatched

        if i < len(validated) - 1:
            next_number = validated[i + 1]["phone"]
            logger.info(
                f"[BATCH] {batch_id}: ⏳ Waiting {WAIT_AFTER_CALL}s before "
                f"call {i+2} → {next_number}"
            )
            progress["status"] = f"waiting_{int(WAIT_AFTER_CALL)}s_before_next"
            await asyncio.sleep(WAIT_AFTER_CALL)
            progress["status"] = "running"

    progress["status"] = "completed"
    progress["current_number"] = None
    progress["retry_count"] = 0
    logger.info(
        f"[BATCH DONE] {batch_id}: All {dispatched}/{len(validated)} calls dispatched ✅"
    )

    asyncio.create_task(_cleanup_progress(batch_id, delay=600))


async def _cleanup_progress(batch_id: str, delay: float = 600):
    await asyncio.sleep(delay)
    batch_progress.pop(batch_id, None)


def get_batch_progress(batch_id: str) -> Optional[dict]:
    return batch_progress.get(batch_id)
