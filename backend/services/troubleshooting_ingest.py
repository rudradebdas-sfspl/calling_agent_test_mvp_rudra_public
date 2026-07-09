"""
Troubleshooting Excel -> structured KB ingestion.

Reads the IT_TROUBLESHOOTING.xlsx workbook and loads two tables:
  - kb_entries   : one row per KB-TS-xxx troubleshooting item (+ embedding)
  - policy_rules : one row per POL-xxx policy rule (+ embedding)

Retrieval later runs a vector search over `content` (a combined text blob) and
the worker's router reads answer_mode / policy_rule_ref to decide behaviour.

Sheet names expected (from the provided workbook):
  - "HNTroubleshooting Entries"
  - "Policy Rules (KB)"
Header matching is tolerant: we normalise header text and match on keywords, so
small wording/case differences won't break ingestion.
"""
from __future__ import annotations

import io
import logging
import re

from openpyxl import load_workbook
from sqlalchemy import text

from backend.database import SessionLocal
from backend.services.embeddings import embed_texts, to_pgvector_literal

log = logging.getLogger("ts_ingest")

ENTRIES_SHEET_HINTS = ("troubleshooting entries", "entries")
POLICY_SHEET_HINTS = ("policy rules", "policy")


# ---------- helpers ----------
def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _pick_sheet(wb, hints):
    for name in wb.sheetnames:
        n = _norm(name)
        if any(h in n for h in hints):
            return name
    return None


def _find_header_row(rows, needed_hint, max_scan=6):
    """Return the index of the row that looks like the header (contains needed_hint)."""
    for i, r in enumerate(rows[:max_scan]):
        joined = _norm(" ".join("" if c is None else str(c) for c in r))
        if needed_hint in joined:
            return i
    return 0


def _col_map(headers, wanted: dict[str, list[str]]) -> dict[str, int]:
    """Map logical field -> column index by matching header keywords.

    `wanted` maps field_name -> list of keyword groups; a header matches if all
    words of any group appear in it. First matching column wins.
    """
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for field, groups in wanted.items():
        for gi, group in enumerate(groups):
            words = group.split()
            for ci, h in enumerate(norm_headers):
                if ci in out.values():
                    continue
                if all(w in h for w in words):
                    out[field] = ci
                    break
            if field in out:
                break
    return out


def _val(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


# ---------- entries ----------
ENTRY_FIELDS = {
    "entry_id":            ["entry id", "id"],
    "category":            ["category"],
    "sub_category":        ["sub category"],
    "title":               ["issue title", "title"],
    "description":         ["issue description", "description"],
    "root_causes":         ["root causes", "why this happens"],
    "keywords":            ["keywords tags", "keywords"],
    "voice_mode":          ["voice mode"],
    "checklist_steps":     ["checklist steps", "checklist"],
    "total_steps":         ["total steps"],
    "escalation_action":   ["escalation action"],
    "source_ref":          ["source ref"],
    "answer_mode":         ["answer mode"],
    "probing_questions":   ["probing questions", "dynamic probing"],
    "diagnostic_commands": ["diagnostic commands"],
    "max_steps":           ["max steps"],
    "policy_boundary":     ["policy boundary", "must not"],
    "sensitive_actions":   ["sensitive actions"],
    "can_resolve":         ["can resolve"],
    "allowed_depth":       ["allowed depth"],
    "data_restrictions":   ["data privacy restrictions", "privacy restrictions"],
    "compliance_notes":    ["compliance notes"],
    "transfer_trigger":    ["human transfer trigger", "transfer trigger"],
    "transfer_after_step": ["transfer after step"],
    "transfer_priority":   ["transfer priority"],
    "transfer_to":         ["transfer to"],
    "info_to_collect":     ["info to collect"],
    "transfer_script":     ["transfer script"],
    "urgency":             ["urgency level", "urgency"],
    "policy_rule_ref":     ["policy rule ref"],
}


def _entry_content(d: dict) -> str:
    """Build the text used for embedding/retrieval — the words a caller might use."""
    parts = [
        d.get("title"), d.get("category"), d.get("sub_category"),
        d.get("description"), d.get("root_causes"), d.get("keywords"),
    ]
    return "\n".join(p for p in parts if p)


# ---------- policy ----------
POLICY_FIELDS = {
    "rule_id":           ["rule id"],
    "policy_area":       ["policy area"],
    "rule_statement":    ["rule plain statement", "rule statement", "rule"],
    "rationale":         ["why rationale", "rationale"],
    "agent_says":        ["what the agent says", "agent says"],
    "answer_mode":       ["answer mode"],
    "applies_to":        ["applies to"],
    "transfer_priority": ["transfer priority"],
    "owner":             ["owner source", "owner"],
}


def _policy_content(d: dict) -> str:
    parts = [d.get("policy_area"), d.get("rule_statement"), d.get("agent_says")]
    return "\n".join(p for p in parts if p)


# ---------- main entrypoint ----------
async def ingest_workbook(agent_id, data: bytes) -> dict:
    """Parse the workbook bytes and (re)load kb_entries + policy_rules for agent_id.

    Returns counts: {"entries": N, "policies": M}.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    entries_sheet = _pick_sheet(wb, ENTRIES_SHEET_HINTS)
    policy_sheet = _pick_sheet(wb, POLICY_SHEET_HINTS)
    if not entries_sheet:
        raise ValueError(
            f"Could not find a troubleshooting-entries sheet. Sheets: {wb.sheetnames}"
        )

    entries = _parse_entries(wb, entries_sheet)
    policies = _parse_policies(wb, policy_sheet) if policy_sheet else []

    if not entries and not policies:
        raise ValueError("No rows parsed — check the sheet headers/format.")

    # Embed all content in one batched pass each.
    if entries:
        vecs = await embed_texts([e["content"] for e in entries], task_type="RETRIEVAL_DOCUMENT")
        for e, v in zip(entries, vecs):
            e["embedding"] = to_pgvector_literal(v)
    if policies:
        pvecs = await embed_texts([p["content"] for p in policies], task_type="RETRIEVAL_DOCUMENT")
        for p, v in zip(policies, pvecs):
            p["embedding"] = to_pgvector_literal(v)

    db = SessionLocal()
    try:
        aid = str(agent_id)
        # Replace previous ingestion for this agent (idempotent re-upload).
        db.execute(text("DELETE FROM kb_entries WHERE agent_id = CAST(:aid AS UUID)"), {"aid": aid})
        db.execute(text("DELETE FROM policy_rules WHERE agent_id = CAST(:aid AS UUID)"), {"aid": aid})

        for e in entries:
            _insert_entry(db, aid, e)
        for p in policies:
            _insert_policy(db, aid, p)

        db.commit()
        log.info("Ingested %d entries, %d policies for agent %s", len(entries), len(policies), aid)
        return {"entries": len(entries), "policies": len(policies)}
    except Exception:
        db.rollback()
        log.exception("ingest failed")
        raise
    finally:
        db.close()


def _canonical_answer_mode(raw) -> str | None:
    """Normalise the many answer-mode string variants to one of two values.

    Returns 'KB-Fetch Only', 'LLM-Answered', or None (row is not a real entry —
    e.g. the descriptor/legend row whose answer_mode starts with 'HOW ...').
    """
    n = _norm(raw)
    if n.startswith("kb fetch"):
        return "KB-Fetch Only"
    if n.startswith("llm answered"):
        return "LLM-Answered"
    return None  # descriptor/instruction/blank -> skip


def _parse_entries(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr_i = _find_header_row(rows, "entry id")
    headers = rows[hdr_i]
    cmap = _col_map(headers, ENTRY_FIELDS)
    out = []
    for r in rows[hdr_i + 1:]:
        eid = _val(r, cmap.get("entry_id"))
        if not eid or not re.match(r"(?i)kb[- ]?ts\s*-?\s*\d", eid):
            continue  # not a KB-TS-#### row
        mode = _canonical_answer_mode(_val(r, cmap.get("answer_mode")))
        if mode is None:
            continue  # descriptor/legend row or missing mode -> skip
        d = {f: _val(r, cmap.get(f)) for f in ENTRY_FIELDS}
        d["answer_mode"] = mode
        d["total_steps"] = _to_int(d.get("total_steps"))
        d["max_steps"] = _to_int(d.get("max_steps"))
        # normalise policy ref like "POL-001" (blank/"none"/"n/a" -> None)
        pr = (d.get("policy_rule_ref") or "").strip()
        d["policy_rule_ref"] = pr if re.match(r"(?i)pol[- ]?\d", pr) else None
        content = _entry_content(d)
        if not content:
            continue
        d["content"] = content
        out.append(d)
    return out


def _parse_policies(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr_i = _find_header_row(rows, "rule id")
    headers = rows[hdr_i]
    cmap = _col_map(headers, POLICY_FIELDS)
    out = []
    for r in rows[hdr_i + 1:]:
        rid = _val(r, cmap.get("rule_id"))
        if not rid or not re.match(r"(?i)pol[- ]?\d", rid):
            continue
        d = {f: _val(r, cmap.get(f)) for f in POLICY_FIELDS}
        content = _policy_content(d)
        if not content:
            continue
        d["content"] = content
        out.append(d)
    return out


_ENTRY_COLS = [
    "entry_id", "category", "sub_category", "title", "description", "root_causes",
    "keywords", "voice_mode", "checklist_steps", "total_steps", "escalation_action",
    "source_ref", "answer_mode", "probing_questions", "diagnostic_commands",
    "max_steps", "policy_boundary", "sensitive_actions", "can_resolve",
    "allowed_depth", "data_restrictions", "compliance_notes", "transfer_trigger",
    "transfer_after_step", "transfer_priority", "transfer_to", "info_to_collect",
    "transfer_script", "urgency", "policy_rule_ref", "content",
]


def _insert_entry(db, aid, e):
    cols = ", ".join(["agent_id"] + _ENTRY_COLS + ["embedding"])
    binds = ["CAST(:agent_id AS UUID)"] + [f":{c}" for c in _ENTRY_COLS] + ["CAST(:embedding AS halfvec)"]
    params = {"agent_id": aid, **{c: e.get(c) for c in _ENTRY_COLS}, "embedding": e.get("embedding")}
    db.execute(text(f"INSERT INTO kb_entries ({cols}) VALUES ({', '.join(binds)})"), params)


_POLICY_COLS = [
    "rule_id", "policy_area", "rule_statement", "rationale", "agent_says",
    "answer_mode", "applies_to", "transfer_priority", "owner", "content",
]


def _insert_policy(db, aid, p):
    cols = ", ".join(["agent_id"] + _POLICY_COLS + ["embedding"])
    binds = ["CAST(:agent_id AS UUID)"] + [f":{c}" for c in _POLICY_COLS] + ["CAST(:embedding AS halfvec)"]
    params = {"agent_id": aid, **{c: p.get(c) for c in _POLICY_COLS}, "embedding": p.get("embedding")}
    db.execute(text(f"INSERT INTO policy_rules ({cols}) VALUES ({', '.join(binds)})"), params)